# -*- coding: utf-8 -*-
import os
from vocab_6th import WEEKS, all_words
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = "/sessions/lucid-gallant-dirac/mnt/youbuddy-challenge-claude"
SRC_LABEL = {"new": "신규", "1기": "재사용·1기", "2기": "재사용·2기"}
SRC_TAG = {"new": "🆕 신규", "1기": "♻️ 1기", "2기": "♻️ 2기"}

# ---------- Markdown ----------
md = []
md.append("# YOUBUDDY 6기 어휘안 (검수용)\n")
md.append("> 총 **60개** = 신규 51 + 1기 재사용 5 + 2기 재사용 4  ·  1~5기 단어와 **중복 0** (자동검증 통과)")
md.append("> 시작 **2026-06-08(월)**, 20 영업일  ·  각 단어는 5기와 동일한 카드 포맷(뜻·예문·유의어·뉘앙스)으로 채워집니다.\n")
md.append("승인해주시면 이 60개를 6th/index.html의 Day1~20에 그대로 주입합니다. 바꾸고 싶은 단어는 'DayN의 X → Y' 식으로 알려주세요.\n")

for w in WEEKS:
    md.append(f"\n## Week {w['n']} · {w['title']} ({w['subtitle']})\n")
    for d in w["days"]:
        md.append(f"\n### Day {d['day']}. {d['title']}\n")
        md.append("| 출처 | 표현 | 품사 | 뜻 | 예문 | 해석 | 유의어 |")
        md.append("|---|---|---|---|---|---|---|")
        for wd in d["words"]:
            syn = ", ".join(wd["syn"])
            ex_en = wd["ex_en"].replace("|", "/")
            ex_kr = wd["ex_kr"].replace("|", "/")
            md.append(f"| {SRC_TAG[wd['src']]} | **{wd['en']}** | {wd['pos']} | {wd['def']} | {ex_en} | {ex_kr} | {syn} |")
        # nuance block
        for wd in d["words"]:
            md.append(f"\n- 🧡 *{wd['en']}*: {wd['nuance']}")
        md.append("")

with open(os.path.join(ROOT, "6기_어휘안_검수.md"), "w", encoding="utf-8") as f:
    f.write("\n".join(md))

# ---------- XLSX ----------
wb = Workbook()
ws = wb.active
ws.title = "6기 어휘 60"

headers = ["Week", "Day", "출처", "영어표현", "품사", "뜻", "예문 (English)", "예문 해석", "유의어", "뉘앙스 (유버디 코멘트)"]
ws.append(headers)

navy = "1F3A5F"; cream = "FFF4E6"; gnew = "E8F0FE"; g1 = "FDECEC"; g2 = "EAF6EC"
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
for c in ws[1]:
    c.font = hfont
    c.fill = PatternFill("solid", fgColor=navy)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 26

srcfill = {"new": gnew, "1기": g1, "2기": g2}
r = 2
for row in all_words():
    ws.append([row["week"], row["day"], SRC_LABEL[row["src"]], row["en"], row["pos"],
               row["def"], row["ex_en"], row["ex_kr"], ", ".join(row["syn"]), row["nuance"]])
    for c in ws[r]:
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(r, 4).font = Font(name="Arial", size=10, bold=True)
    fill = PatternFill("solid", fgColor=srcfill[row["src"]])
    ws.cell(r, 3).fill = fill
    ws.cell(r, 3).alignment = Alignment(horizontal="center", vertical="top")
    r += 1

widths = [6, 6, 11, 22, 14, 26, 34, 30, 26, 50]
for i, wdt in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = wdt
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{r-1}"

wb.save(os.path.join(ROOT, "6기_어휘안_검수.xlsx"))
print("생성 완료:", r-2, "행")
