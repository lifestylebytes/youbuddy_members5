"""Create the 5기 어휘 템플릿.xlsx file for Korean business English learning app."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import os
# Resolve output path for both macOS host and Linux workspace mount
_CANDIDATE_DIRS = [
    "/Users/gyo/Downloads/youbuddy-challenge-claude",
    "/sessions/compassionate-quirky-hypatia/mnt/youbuddy-challenge-claude",
]
_OUT_DIR = next((d for d in _CANDIDATE_DIRS if os.path.isdir(d)), _CANDIDATE_DIRS[0])
OUTPUT_PATH = os.path.join(_OUT_DIR, "5기_어휘_템플릿.xlsx")

# ---------- Weekly / daily structure ----------
WEEKS = [
    {
        "week_n": 1,
        "week_title": "Onboarding",
        "week_subtitle": "방향성 잡고 바로 시작",
        "days": [1, 2, 3, 4, 5],
    },
    {
        "week_n": 2,
        "week_title": "Meetings & Messaging",
        "week_subtitle": "일상 회의 영어 루틴 완성",
        "days": [6, 7, 8, 9, 10],
    },
    {
        "week_n": 3,
        "week_title": "Stakeholder Work",
        "week_subtitle": "이해관계자와의 대화 능숙하게",
        "days": [11, 12, 13, 14, 15],
    },
    {
        "week_n": 4,
        "week_title": "Leadership & Wrap-up",
        "week_subtitle": "리더십 표현 + 4주 마무리",
        "days": [16, 17, 18, 19, 20],
    },
]

HEADERS = [
    "week_n", "week_title", "week_subtitle",
    "day", "day_title", "goal_en", "goal_kr",
    "idx", "en", "pos", "def",
    "ex_en", "ex_kr",
    "syn1", "syn2", "syn3",
    "nuance",
]

COLUMN_WIDTHS = {
    "A": 8, "B": 14, "C": 26,
    "D": 8, "E": 26, "F": 40, "G": 26,
    "H": 8, "I": 18, "J": 14, "K": 18,
    "L": 40, "M": 34,
    "N": 14, "O": 14, "P": 14,
    "Q": 40,
}

HEADER_FILL = PatternFill(start_color="FFE8D6", end_color="FFE8D6", fill_type="solid")
EN_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
ARIAL = "Arial"


def build_vocab_sheet(ws):
    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name=ARIAL, bold=True, size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Data rows — 60 rows total, 3 per day × 20 days
    row = 2
    for week in WEEKS:
        for day in week["days"]:
            for idx in range(3):
                ws.cell(row=row, column=1, value=week["week_n"])
                ws.cell(row=row, column=2, value=week["week_title"])
                ws.cell(row=row, column=3, value=week["week_subtitle"])
                ws.cell(row=row, column=4, value=day)
                # Column E (day_title), F (goal_en), G (goal_kr) blank
                ws.cell(row=row, column=8, value=idx)
                # Columns I-Q blank

                # Apply font + alignment to all 17 columns in this row
                for col in range(1, 18):
                    c = ws.cell(row=row, column=col)
                    c.font = Font(name=ARIAL, size=10)
                    c.alignment = Alignment(vertical="top", wrap_text=True)

                # Highlight 'en' column (I = 9) as required
                ws.cell(row=row, column=9).fill = EN_FILL

                row += 1

    # Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


def build_guide_sheet(ws):
    rows = [
        ("5기 어휘 입력 가이드", {"bold": True, "size": 14}),
        ("", {}),
        ("📋 꼭 지켜주세요", {"bold": True}),
        ("· 총 60행 (20일 × 3표현). 빠지거나 남는 행이 있으면 안 됩니다.", {}),
        ("· idx 컬럼은 0, 1, 2 순서 고정 (같은 day 내 표현 순서).", {}),
        ("· 'en' 컬럼의 표기가 단어 시험 정답 판정에 그대로 쓰입니다. 대소문자·따옴표·하이픈까지 정확히 써주세요.", {}),
        ("· day_title / goal_en / goal_kr 은 각 day의 첫 행(idx=0)에만 쓰면 됩니다. 나머지 행은 비워두세요.", {}),
        ("", {}),
        ("🗂️ 컬럼 설명", {"bold": True}),
        ("week_n: 주차 번호 (1~4). 이미 채워져 있어요.", {}),
        ("week_title: 주차 영문 제목. 이미 채워져 있어요.", {}),
        ("week_subtitle: 주차 한국어 부제. 이미 채워져 있어요.", {}),
        ("day: 일차 번호 (1~20). 이미 채워져 있어요.", {}),
        ("day_title: 일차 한국어 부제 (예: '목표/방향성 + 킥오프').", {}),
        ('goal_en: 일차 영어 목표 문장 (예: "Align on our North Star, then kick off.").', {}),
        ("goal_kr: 일차 한국어 목표 설명 (예: '4주의 방향성을 먼저 정리하고 출발해요.').", {}),
        ("idx: 같은 day 내 표현 순번 (0, 1, 2). 이미 채워져 있어요.", {}),
        ("en: 영어 표현. 예) 'align on', 'kick off', 'North Star'", {}),
        ("pos: 품사. noun / verb / phrase / phrasal verb / acronym / idiom 중 택1", {}),
        ("def: 짧은 한국어 뜻 (10자 내외).", {}),
        ("ex_en: 예문 영어 (표현 포함된 자연스러운 한 문장).", {}),
        ("ex_kr: 예문 한국어 — 이 문장이 앱에서 '이런 상황, 내 업무로 바꿔 쓴다면?' 박스에 그대로 쓰입니다. 반드시 업무 장면 위주로.", {}),
        ("syn1, syn2, syn3: 유의어 2~3개. syn3만 비워도 괜찮습니다.", {}),
        ("nuance: 대표 표현 뉘앙스 (한국어 1~2문장). 이 표현이 \"어떤 상황에서 어떤 느낌으로 쓰이는지\" 설명. 예) '여러 목표 중 \"절대 흔들리면 안 되는 것\"을 가리킬 때 써요.'", {}),
        ("", {}),
        ("✅ 체크리스트", {"bold": True}),
        ("· 60행 전부 I~Q 컬럼 채웠는가?", {}),
        ("· 각 day의 첫 행에 day_title / goal_en / goal_kr 있는가?", {}),
        ("· en 표기가 정확한가? (정답 판정에 사용됨)", {}),
        ("· ex_kr이 실제 업무 상황에 맞는가?", {}),
    ]

    for i, (text, style) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(
            name=ARIAL,
            size=style.get("size", 10),
            bold=style.get("bold", False),
        )
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 90


def main():
    wb = Workbook()
    ws_vocab = wb.active
    ws_vocab.title = "어휘"
    build_vocab_sheet(ws_vocab)

    ws_guide = wb.create_sheet("가이드")
    build_guide_sheet(ws_guide)

    wb.save(OUTPUT_PATH)

    # ---------- Verification ----------
    from openpyxl import load_workbook
    wb2 = load_workbook(OUTPUT_PATH)
    assert "어휘" in wb2.sheetnames, "Sheet '어휘' missing"
    assert "가이드" in wb2.sheetnames, "Sheet '가이드' missing"

    vs = wb2["어휘"]
    assert vs.max_row == 61, f"Expected 61 rows, got {vs.max_row}"
    assert vs.max_column == 17, f"Expected 17 columns, got {vs.max_column}"

    # Verify pre-population of key structural columns
    expected_structure = []
    for week in WEEKS:
        for day in week["days"]:
            for idx in range(3):
                expected_structure.append(
                    (week["week_n"], week["week_title"], week["week_subtitle"], day, idx)
                )
    assert len(expected_structure) == 60

    for i, (wn, wt, wst, d, idx) in enumerate(expected_structure, start=2):
        assert vs.cell(row=i, column=1).value == wn, f"row {i} week_n"
        assert vs.cell(row=i, column=2).value == wt, f"row {i} week_title"
        assert vs.cell(row=i, column=3).value == wst, f"row {i} week_subtitle"
        assert vs.cell(row=i, column=4).value == d, f"row {i} day"
        assert vs.cell(row=i, column=8).value == idx, f"row {i} idx"
        # E, F, G, I-Q should be blank
        for blank_col in [5, 6, 7, 9, 10, 11, 12, 13, 14, 15, 16, 17]:
            assert vs.cell(row=i, column=blank_col).value in (None, ""), (
                f"row {i} col {blank_col} should be blank"
            )

    # Check freeze panes
    assert vs.freeze_panes == "A2", f"freeze_panes is {vs.freeze_panes}"

    # Check column widths
    for col, w in COLUMN_WIDTHS.items():
        actual = vs.column_dimensions[col].width
        assert actual == w, f"col {col} width {actual} != {w}"

    # Check header row values
    for col_idx, header in enumerate(HEADERS, start=1):
        assert vs.cell(row=1, column=col_idx).value == header

    # Spot-check that idx=0 rows are at the expected positions
    idx0_rows = [2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35, 38, 41, 44, 47, 50, 53, 56, 59]
    for r in idx0_rows:
        assert vs.cell(row=r, column=8).value == 0, f"row {r} should have idx=0"

    print("OK - File created and verified.")
    print(f"Path: {OUTPUT_PATH}")
    print(f"어휘 sheet: {vs.max_row} rows, {vs.max_column} columns")
    print(f"가이드 sheet: {wb2['가이드'].max_row} rows")


if __name__ == "__main__":
    main()
